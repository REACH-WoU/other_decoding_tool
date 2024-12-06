from sklearn.metrics.pairwise import cosine_similarity
import pandas as pd
import numpy as np
from openai import OpenAI
import zipfile
import py7zr
import rarfile
import openpyxl
import os, stat, shutil
import tempfile

def get_embedding(df, column, model):
    """
    Generate embeddings for unique values in a specified column using a given model.
    
    Parameters:
        df (pd.DataFrame): The input DataFrame containing the column to embed.
        column (str): Name of the column for which embeddings will be generated.
        model (SentenceTransformer): Pre-trained embedding model.

    Returns:
        pd.DataFrame: The input DataFrame merged with a new column of embeddings for the specified column.
    """
    # Extract unique values
    unique_values = df[column].unique().tolist()
     # Clean and preprocess unique values by removing digits and special characters
    unique_values_clean = [value.replace("0", "").replace("1", "").\
                    replace("2", "").replace("3", "").replace("4", "").\
                    replace("5", "").replace("6", "").replace("7", "").\
                    replace("8", "").replace("9", "").replace("_", "") for value in unique_values]
    unique_values_clean = [value.lower().strip().replace(", ", " , ") for value in unique_values]
    # Generate embeddings for the cleaned unique values
    embeddings = list(model.encode(unique_values_clean))
    short_df = pd.DataFrame({"value": unique_values, "embedding": embeddings})
    # Join input frame with recoded frame of unique values with left join
    return pd.merge(df, short_df, left_on=column, right_on="value", how="left")


def get_openai_embedding(df, column):
    """
    Generate embeddings for unique values in a DataFrame column using the OpenAI API.
    Parameters:
        df (pd.DataFrame): The input DataFrame containing the column to embed.
        column (str): The column name for which embeddings will be generated.

    Returns:
        pd.DataFrame: The input DataFrame with an additional column for embeddings.
    """
    client = OpenAI(api_key="API_KEY")
    unique_values = df[column].unique()
    unique_values = [value.lower().strip().replace(", ", " , ") for value in unique_values]
    # Filter out empty strings
    unique_values = [value for value in unique_values if value != '']

    embeddings = []
    # Batch process embeddings to handle large datasets
    for i in range(0, len(unique_values), 100):
        try:
            res = client.embeddings.create(input = unique_values[i:i + 100], model="text-embedding-ada-002").data
        except Exception as e:
            print(unique_values[i:i + 100])
            raise e
        embeddings += [r.embedding for r in res]
    
    # Handle empty values, append for them zero vector
    unique_values.append('')
    embeddings.append([0] * len(embeddings[0]))

    # Join input frame with recoded frame of unique values with left join
    short_df = pd.DataFrame({"value": unique_values, "embedding": embeddings})
    return pd.merge(df, short_df, left_on=column, right_on="value", how="left")


def extract_archive(archive_path, archive_dir):
    """
    Extract the contents of an archive file (.zip, .7z, .rar) into a specified directory.
    
    Parameters:
        archive_path (str): Path to the archive file.
        archive_dir (str): Directory where the contents will be extracted.

    Raises:
        ValueError: If the archive format is unsupported.
    """
    temp_dir = tempfile.mkdtemp()

    try:
        # Join input frame with recoded frame of unique values with left join
        if archive_path.endswith('.zip'):
            with zipfile.ZipFile(archive_path, 'r') as archive:
                archive.extractall(temp_dir)
        elif archive_path.endswith('.7z'):
            with py7zr.SevenZipFile(archive_path, mode='r') as archive:
                archive.extractall(path=temp_dir)
        elif archive_path.endswith('.rar'):
            with rarfile.RarFile(archive_path, 'r') as archive:
                archive.extractall(temp_dir)
        # elif archive_path.endswith('.tar'):
        #     with tarfile.open(archive_path, 'r') as archive:
        #         archive.extractall(path=temp_dir)
        else:
            raise ValueError("Unsupported archive format")

        # Move extracted contents to the "archive_dir" directory
        extracted_items = os.listdir(temp_dir)
        if len(extracted_items) == 1 and os.path.isdir(os.path.join(temp_dir, extracted_items[0])):
            extracted_dir = os.path.join(temp_dir, extracted_items[0])
            for item in os.listdir(extracted_dir):
                shutil.move(os.path.join(extracted_dir, item), archive_dir)
        else:
            for item in extracted_items:
                shutil.move(os.path.join(temp_dir, item), archive_dir)
    finally:
        shutil.rmtree(temp_dir)


def preprocess_data(df, isOther=False):
    """
    Preprocess a DataFrame by renaming columns for consistency and removing NaN values.
    
    Parameters:
        df (pd.DataFrame): Input DataFrame to preprocess.
        isOther (bool): If True, special preprocessing for "Other" data is applied.
    
    Returns:
        pd.DataFrame: Preprocessed DataFrame.
    """
    df.rename(columns={"INVALID other (insert yes or leave blank)": "invalid"}, inplace=True)
    df.rename(columns={"EXISTING other (copy the exact wording from the options in column choices.label)": "existing"}, inplace=True)
    df.rename(columns={"TRUE other (provide a better translation if necessary)": "translation"}, inplace=True)
    if not isOther and "response.en" in df.columns:
        df.dropna(subset=["response.en"], inplace=True)
    
    df.reset_index(drop=True, inplace=True)

    return df

def preprocess_data_longit(df, isOther=False):
    """
    Preprocess longitudinal data by renaming and resetting DataFrame structure.
    
    Parameters:
        df (pd.DataFrame): Input DataFrame to preprocess.
        isOther (bool): If True, special preprocessing for "Other" data is applied.
    
    Returns:
        pd.DataFrame: Preprocessed longitudinal DataFrame.
    """
    df.rename(columns={"INVALID other (insert yes or leave blank)": "invalid"}, inplace=True)
    df.rename(columns={"EXISTING other (copy the exact wording from the options in column choices.label)": "existing"}, inplace=True)
    df.rename(columns={"TRUE other (provide a better translation if necessary)": "translation"}, inplace=True)
    df.rename(columns={"response.en.from.uk": "response.en"}, inplace=True)
    if not isOther and "response.en" in df.columns:
        df.dropna(subset=["response.en"], inplace=True)
    
    df.reset_index(drop=True, inplace=True)

    return df

def check_basic_columns(df):
    """
    Check if the required columns exist in the DataFrame.
    
    Parameters:
        df (pd.DataFrame): Input DataFrame to check.
    
    Returns:
        str or None: Error message if a required column is missing, otherwise None.
    """
    if "response.en" not in df.columns:
        return"response.en column not found"
    if "invalid" not in df.columns:
        return "invalid column not found"
    if "existing" not in df.columns:
        return "existing column not found"
    if "translation" not in df.columns:
        return "translation column not found"

def get_max_similarity(vector, keys):
    """
    Compute the maximum cosine similarity between a vector and a set of keys.
    
    Parameters:
        vector (list): The input vector to compare.
        keys (list): A list of vectors to compute similarities with.
    
    Returns:
        tuple: Maximum similarity value and the index of the most similar vector.
    """
    similarities = cosine_similarity([vector], keys)
    max_similarity = np.max(similarities)
    max_similarity_index = np.argmax(similarities)
    return max_similarity, max_similarity_index


def get_max_similarity_index(similarities):
    """
    Find the index of the maximum similarity value in a similarity matrix.
    
    Parameters:
        similarities (list): A list of similarity values.
    
    Returns:
        tuple: Maximum similarity value and its corresponding index.
    """
    max_index = np.argmax(similarities)
    max_similarity = similarities[max_index]
    return max_similarity, max_index


def get_same_format(other_file_path, others):
    """
    Create an output Excel file with the same format as the input file,
    populating it with the recoded data.
    
    Parameters:
        other_file_path (str): Path to the original "Other" file.
        others (pd.DataFrame): DataFrame containing the recoded data.
    """
    # Copy the original file structure to the output directory
    shutil.copyfile(other_file_path, "output_dir/output.xlsx")
    
    # Load the copied workbook
    workbook = openpyxl.load_workbook("output_dir/output.xlsx")
    sheet = workbook.active

    # Clear existing data (starting from the second row)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.value = None

    # Populate the sheet with new data
    for r_idx, row in enumerate(others.values, 2):
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row=r_idx, column=c_idx, value=value)
    
    # Save the updated file
    workbook.save("output_dir/output.xlsx")

def check_file_correct(other_file_path):
    """
    Validate that the provided file can be opened and is in a correct Excel format and with edit permissions
    
    Parameters:
        other_file_path (str): Path to the file to check.
    
    Returns:
        bool: True if the file is valid, False otherwise.
    """
    try:
        # Try to load the workbook to ensure it is a valid Excel file
        workbook = openpyxl.load_workbook(other_file_path)
        return True
    except Exception as e:
        print(e)
        return False
